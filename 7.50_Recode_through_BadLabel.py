
###############################################################################################
###############################################################################################

# Name:             7.50_Recode_through_BadLabel.py
# Author:           Kelly Meehan, USBR
# Created:          20180618
# Updated:          20200825 
# Version:          Created using Python 3.6.8 

# Requires:         ArcGIS Pro 

# Notes:            This script is intended to be used for a Script Tool within ArcGIS Pro; it is not intended as a stand-alone script.

# Description:      This tool generates an output similar to that of ERDAS IMAGINE's Save Zonal Statistics To Polygon Attributes tool;
#                   it generates an Excel spreadsheet with a normalized difference index of badness for each signature.

#----------------------------------------------------------------------------------------------

# Tool setup:       The script tool's properties can be set as follows: 
#
#                      Parameters tab:    
#                           Classified Raster               Raster Layer (Data Type) > Required (Type) > Input (Direction)  
#                           Field Borders Feature Class     Feature Layer (Data Type) > Required (Type) > Input (Direction)
#                           Image Directory                 Workspace (Data Type) > Required (Type) > Input (Direction)                    
#                           Shapefile Directory             Workspace (Data Type) > Required (Type) > Input (Direction)                    
#                           Documents Directory             Workspace (Data Type) > Required (Type) > Input (Direction)                    
#                           Geodatabase                     Workspace (Data Type) > Required (Type) > Input (Direction)
#                           Iteration Number                String (Data Type) > Required (Type) > Input (Direction)
#
#                       Validation tab:
#

# import arcpy, re

# class ToolValidator(object):
#     """Class for validating a tool's parameter values and controlling
#     the behavior of the tool's dialog."""

#     def __init__(self):
#         """Setup arcpy and the list of tool parameters.""" 
#         self.params = arcpy.GetParameterInfo()

#     def initializeParameters(self):
#         """Refine the properties of a tool's parameters. This method is 
#         called when the tool is opened."""

#     def updateParameters(self):
#         """Modify the values and properties of parameters before internal
#         validation is performed. This method is called whenever a parameter
#         has been changed."""
        
#         # Set a default value for Edited Field Borders Shapefile
#         if self.params[0].value:
#             if not self.params[1].altered:
#                 raster_name = os.path.basename(self.params[0].value.value)
#                 region_time_caps = raster_name.rsplit(sep = '_', maxsplit = 2)[0].upper()
#                 img_directory = os.path.dirname(self.params[0].value.value)
#                 covs_directory = os.path.abspath(os.path.join(img_directory, '..', 'covs_' + region_time_caps))
#                 edtd_shp = os.path.join(covs_directory, region_time_caps + '_edtd.shp')
#                 iteration = re.split('[_.]', os.path.basename(raster_name))[4]
#                 self.params[1].value = edtd_shp  
                
#         # Set a default value for Iteration Number
#             if not self.params[6].altered: 
#                 self.params[6].value = iteration
                     
#         # Set a default value for Image Directory
#         if self.params[1].value:
#             if not self.params[2].altered:
#                 self.params[2].value = img_directory        

#         # Set a default value for Shapefile Directory
#             if not self.params[3].altered:
#                 self.params[3].value = covs_directory

#         # Set a default value for Documents Directory
#             if not self.params[4].altered:
#                 docs_directory = os.path.abspath(os.path.join(covs_directory, '..', 'docs_' + region_time_caps))
#                 self.params[4].value = docs_directory

#         # Set a default value for Project Geodatabase
#             if not self.params[5].altered:
#                 gdb_directory = os.path.abspath(os.path.join(covs_directory, '..', region_time_caps + '.gdb'))
#                 self.params[5].value = gdb_directory
                             
#     def updateMessages(self):
#         """Modify the messages created by internal validation for each tool
#         parameter. This method is called after internal validation."""
        
#         # Ensure that Iteration Number is a two digit value
#         if self.params[6].value:
#             iteration_value = self.params[6].value
#             if len(iteration_value) != 2 or not iteration_value.isdigit():
#                 self.params[6].setWarningMessage('{0} is not an appropriate value to pass to Iteration Number parameter. Please provide a two digit integer (e.g. for first iteration input 01)'.format(iteration_value))

#     def isLicensed(self):
#         """Set whether tool is licensed to execute."""
#         return True
    
###############################################################################################
###############################################################################################

# This script will:

# 0. Set-up
# 1. Make a copy of Classified Raster and save as Classified Tiff
# 2. Delete attribute table field: Crop if already present in Classified Tiff
# 3. Add attribute table field, Crop, to Classified Tiff
# 4. Populate attribute table field, Crop, with extracted CROP_TYPE value from concatenated Signame 
# 5. Reclassify Classified Tiff based on attribute table field, Crop
# 6. In Edited Field Borders Shapefile attribute table field, MAJORITY, assign each field a classification value based on what the majority of pixels were assigned to  
# 7. Create Label Shapefile, an exact copy of Edited Field Borders Shapefile
# 8. Create two new fields: Crop_label and crop to Label Shapefile   
# 9. Populate Label Shapefile attribute table fields, Crop_label and crop, to values of MAJORITY and CROP_TYPE, respectively
# 10. Overwrite Label Shapefile attribute table field, Crop_label with known value, crop (i.e. CROP_TYPE) for known static ground truth fields ('CLASS' == 2)
# 11. Create Label Frequency Table with sum acreage for fields aggregated by the unique combination of : CLASS, aa, crop, and Crop_label
# 12. Convert Label Frequency Table to Excel table
# 13. Rename MAJORITY attribute table field name to MAJORITY** in Edited Field Borders Shapefile, allowing subsequent iterations to reuse the name MAJORITY
# 14. Create Bad Label Shapefile representing training fields misclassified
# 15. Generate Label Shapefile that were training fields, regardless of whether classified correctly or not
# 16. Cross-tabulate area of misclassified pixels by Signame and CROP_TYPE
# 17. Generate Bad Signame Excel Workbook, a pivoted and manipulated version of Bad Signame Geodatabase Table

#----------------------------------------------------------------------------------------------

# 0. Set-up

# 0.0 Install necessary packages

import arcpy, os, pandas, re, sys, psutil, time
from arcpy.sa import RemapValue, Reclassify, ZonalStatisticsAsTable, TabulateArea 

#--------------------------------------------

# 0.1 Read in tool parameters

# User selects Classified Raster (ERDAS IMAGINE .img file)
classified_raster = arcpy.GetParameterAsText(0) 

# User selects Edited Field Borders Shapefile
edited_field_borders_shapefile = arcpy.GetParameterAsText(1)

# User selects Image Directory
img_path = arcpy.GetParameterAsText(2)

# User selects Coverage Directory
covs_path = arcpy.GetParameterAsText(3)

# User selects Documents Directory
docs_path = arcpy.GetParameterAsText(4)

# User selects Project Geodatabase
gdb_path = arcpy.GetParameterAsText(5)

# User selects two digit classification iteration number
iteration_number = arcpy.GetParameterAsText(6)

#--------------------------------------------

# 0.2 Set environment settings

# Overwrite output
arcpy.env.overwriteOutput = True

# Set mask
arcpy.env.mask = edited_field_borders_shapefile

# Set snap raster
arcpy.env.snapRaster = classified_raster

# 0.3 Check out Spatial Analyst Extension
arcpy.CheckOutExtension('Spatial')

#----------------------------------------------------------------------------------------------

# 1. Make a copy of Classified Raster and save as Classified Tiff (so as to avoid in later steps error 00499: table is not editable)

classified_tiff = os.path.splitext(classified_raster)[0] + '.tif'

arcpy.CopyRaster_management(in_raster = classified_raster, out_rasterdataset = classified_tiff)

arcpy.AddMessage('Generated Classified Tiff: ' + classified_tiff)

#----------------------------------------------------------------------------------------------

# 2. Delete attribute table field, Crop, if already present in Classified Tiff (ignoring case)

tiff_fields = [field.name.upper() for field in arcpy.ListFields(dataset = classified_tiff)]
for i in tiff_fields:
    if i == 'CROP':
        arcpy.DeleteField_management(in_table = classified_tiff, drop_field = i)

#----------------------------------------------------------------------------------------------

# 3. Add attribute table field, Crop, to Classified Tiff

arcpy.AddField_management(in_table = classified_tiff, field_name = 'Crop', field_type = 'SHORT')

arcpy.AddMessage('Added attribute table field: Crop to Classified Tiff')

#----------------------------------------------------------------------------------------------

# 4. Populate attribute table field, Crop, with extracted crop type value from concatenated Signame 

input_table_fields = ['Class_Name', 'Crop']
with arcpy.da.UpdateCursor(in_table = classified_tiff, field_names = input_table_fields) as cursor:
    for row in cursor:
        if row[0] == 'Unclassified':
            row[1] = 'Unclassified'
        elif row[0] == ' ':
           row[1] = ''
        else:
            row[1] = row[0].split('-')[2] 
            cursor.updateRow(row)
            
arcpy.AddMessage(''''Populated Classified Tiff's attribute table field, Crop, with crop code value extracted from Signame''')

#----------------------------------------------------------------------------------------------

# 5. Reclassify Classified Tiff based on attribute table field, Crop

region_and_time = os.path.basename(edited_field_borders_shapefile).rsplit(sep = '_', maxsplit = 1)[0]
region_and_time_caps = region_and_time.upper()
reclassified_raster_name = region_and_time + '_reclassified_' + iteration_number + '.img'
reclassified_raster = os.path.join(img_path, reclassified_raster_name)

remap_values = RemapValue([[0,'NODATA']])

if 'out_reclassify' in locals():
    arcpy.AddMessage('delecting out_reclassify')
    del out_reclassify
    
out_reclassify = Reclassify(in_raster = classified_tiff, reclass_field = 'Crop', remap = remap_values, missing_values = 'DATA') # NOTE: Passing argument 'DATA' to missing_values parameter allows you to leave all values other than 0 as they are.

# Test whether Reclassified Raster can be generated (i.e. if re-running tool it may not be able to delete pre-exising raster) and exit script if not  
try:
    out_reclassify.save(reclassified_raster)
except RuntimeError:
    arcpy.AddError('Cannot overwrite pre-existing ' + reclassified_raster + '; please re-run tool once more')
    sys.exit(0)

arcpy.AddMessage('Generated Reclassified Raster based on attribute table field, Crop')

# Delete output reclassified raster variable otherwise it will cause a lock (ERROR 000871: Unable to delete the output) if the tool is re-run for the same iteration
del out_reclassify

#----------------------------------------------------------------------------------------------

# 6. In Edited Field Borders Shapefile attribute table field, MAJORITY, assign each field a classification value based on what the majority of pixels were assigned to  

# Create Zonal Statistics Majority Table 

edited_field_borders_shapefile_name = os.path.basename(edited_field_borders_shapefile)
majority_table_name = region_and_time_caps + '_majority_' + iteration_number + '.dbf'
majority_table = os.path.join(docs_path, majority_table_name)

ZonalStatisticsAsTable(in_zone_data = edited_field_borders_shapefile, zone_field = 'FIELD_ID', in_value_raster = reclassified_raster, out_table = majority_table, ignore_nodata = 'DATA', statistics_type = 'MAJORITY')

arcpy.AddMessage('Generated Zonal Statistics Majority Table: ' + majority_table)

# Delete attribute table fields: MAJORITY or MAJORITY** (where ** corresponds to this iteration) in case user needs to re-run this same iteration

delete_fields_list = ['MAJORITY', 'MAJORITY' + iteration_number]
for d in delete_fields_list:
    if arcpy.ListFields(dataset = edited_field_borders_shapefile, wild_card = d):
        arcpy.DeleteField_management(in_table = edited_field_borders_shapefile, drop_field = d)
        arcpy.AddMessage('Deleting pre-existing field from Edited Field Borders Shapefile: ' + str(d)) 

# Join the majority values from the output table of Zonal Statistics as Table to the Edited Field Borders Shapefile

arcpy.JoinField_management(in_data = edited_field_borders_shapefile, in_field = 'FIELD_ID', join_table = majority_table, join_field  = 'FIELD_ID', fields = 'MAJORITY')

arcpy.AddMessage('Joined MAJORITY field from Zonal Statistics Table to Edited Field Borders Shapefile') 

#----------------------------------------------------------------------------------------------

# 7. Create Label Shapefile, an exact copy of Edited Field Borders Shapefile

label_shapefile_name = region_and_time_caps + '_label_' + iteration_number + '.shp'
label_shapefile = os.path.join(covs_path, label_shapefile_name)

arcpy.Copy_management(in_data = edited_field_borders_shapefile, out_data = label_shapefile)

arcpy.AddMessage('Created Label Shapefile: ' + label_shapefile)

#----------------------------------------------------------------------------------------------

# 8. Create two new attribute table fields to Label Shapefile, Crop_label and crop 

add_fields_list = ['Crop_label', 'crop']
for a in add_fields_list:
    if arcpy.ListFields(dataset = label_shapefile, wild_card = a):
        arcpy.DeleteField_management(in_table = label_shapefile, drop_field = a)
    arcpy.AddField_management(in_table = label_shapefile, field_name = a, field_type = 'SHORT')

arcpy.AddMessage('Added attribute table fields: Crop_label and crop to Label Shapefile')

#----------------------------------------------------------------------------------------------

# 9. Populate Label Shapefile attribute table fields, Crop_label and crop, with values from MAJORITY and CROP_TYPE, respectively

comparison_fields = ['Crop_label', 'MAJORITY', 'crop', 'CROP_TYPE', 'CLASS']
with arcpy.da.UpdateCursor(label_shapefile, comparison_fields) as cursor:
    for row in cursor:
        row[0] = row[1]/100
        row[2] = row[3]/100
        cursor.updateRow(row)

#----------------------------------------------------------------------------------------------

# 10. Overwrite Label Shapefile attribute table field, Crop_label with known value, crop (i.e. CROP_TYPE) for known static ground truth fields ('CLASS' == 2
        if row[4] == 2:
            row[0] = row[2]
            cursor.updateRow(row)

arcpy.AddMessage('Populated Crop_label and crop attribute table values within Label Shapefile to match MAJORITY and CROP_TYPE and burned in CROP_TYPE for known static fields')

#----------------------------------------------------------------------------------------------

# 11. Create Label Frequency Table with sum acreage for fields aggregated by the unique combination of: CLASS, aa, crop, and Crop_label

frequency_table_name = region_and_time_caps + '_label_' + iteration_number + '_fre.dbf'
frequency_table = os.path.join(docs_path, frequency_table_name)

arcpy.Frequency_analysis(in_table = label_shapefile, out_table = frequency_table, frequency_fields = ['CLASS', 'aa', 'crop', 'Crop_label'], summary_fields = 'ACRES')

arcpy.AddMessage('Created Label Frequency Table')

#----------------------------------------------------------------------------------------------

# 12. Convert Label Frequency Table to Excel file

frequency_table_xlsx_name = os.path.splitext(frequency_table_name)[0] + '.xlsx'
frequency_table_xlsx = os.path.join(docs_path, frequency_table_xlsx_name)

arcpy.TableToExcel_conversion(Input_Table = frequency_table, Output_Excel_File = frequency_table_xlsx)

arcpy.AddMessage('Coverted Label Frequency Table to Excel file')

#----------------------------------------------------------------------------------------------

# 13. Rename MAJORITY attribute table field name to MAJORITY** in Edited Field Borders Shapefile, allowing subsequent iterations to reuse the name MAJORITY

# Used workaround of creating new empty field with desired name, copy values, delete old field since cannot use arcpy.AlterField_management (as it only works on filegeodatabase shapefile) 

field_borders_fields  = [field.name for field in arcpy.ListFields(dataset = edited_field_borders_shapefile)]

for field in field_borders_fields:
    if field == 'MAJORITY':
        arcpy.AddField_management(in_table = edited_field_borders_shapefile, field_name = 'MAJORITY' + iteration_number, field_type = 'SHORT')
        arcpy.CalculateField_management(in_table = edited_field_borders_shapefile, field = 'MAJORITY' + iteration_number, expression = "!MAJORITY!", expression_type = 'PYTHON3')
        arcpy.DeleteField_management(in_table = edited_field_borders_shapefile, drop_field = 'MAJORITY')

#----------------------------------------------------------------------------------------------

# 14. Create Bad Label Shapefile representing training fields misclassified

bad_shapefile_name = region_and_time_caps + '_bad_label_' + iteration_number + '.shp'
bad_shapefile = os.path.join(covs_path, bad_shapefile_name)

arcpy.Select_analysis(in_features = label_shapefile, out_feature_class = bad_shapefile, where_clause = "\"crop\" > 0 AND \"crop\" <> \"Crop_label\" AND \"aa\" = 1")

arcpy.AddMessage('Created Bad Label Shapefile')

#----------------------------------------------------------------------------------------------

# 15. Generate training Label Shapefile, a subset of Label Shapefile that were training fields, regardless of whether classified correctly or not

training_label_shapefile_name = region_and_time_caps + '_training_label_' + iteration_number + '.shp'
training_label_shapefile = os.path.join(covs_path, training_label_shapefile_name)

arcpy.Select_analysis(in_features = label_shapefile, out_feature_class = training_label_shapefile, where_clause = "\"crop\" > 0 AND \"aa\" = 1")

arcpy.AddMessage('Created Training Label Shapefile')

#----------------------------------------------------------------------------------------------

# 16. Cross-tabulate area of misclassified pixels by Signame and CROP_TYPE

# Create Bad Signame Geodatabase Table, a cross-tabulation of area for each misclassified CROP_TYPE by responsible Signame
#   Each row represents misclassified crop types (found within Bad Label Shapefile)
#   Each column represents the corresponding signame that was responsible for the wrong label
#       Note that the column headers read VALUE_*, which correspond to an actual Signame within Classified Tiff; the prefix Value is given as a geodatabase table cannot start with a number and tabulate area runs off of Value 
#   Values within the table are total area in meters of misclassified training fields broken up CROP_TYPE and a value (corresponding to a Signame within Classified Tiff)

bad_sig_table_name = region_and_time_caps + '_bad_signames_' + iteration_number
bad_sig_table = os.path.join(gdb_path, bad_sig_table_name)

# Try to generate Bad Signame Geodatabase Table
try:
    TabulateArea(in_zone_data = bad_shapefile, zone_field = 'crop', in_class_data = classified_tiff, class_field = 'Value', out_table = bad_sig_table)

# If an exception is raised (Execute Error: ERROR 010151: No features found) catch and handle it by giving warning and proceeding

except arcpy.ExecuteError:
    e = sys.exc_info()[1]
    arcpy.AddWarning(e.args[0])
    result = arcpy.GetCount_management(bad_shapefile)
    if int(result[0]) > 0:
        pass
    else:
        arcpy.AddWarning('Bad Label Shapefile has no features and so Bad Signame Geodatabase Table was not generated. Continuing...')

# Create Training Label Signame Geodatabase Table, a cross-tabulation of area for Signame for all training fields regardless of whether they were misclassified or not
#   Rows, columns, and values represent the same as in Bad Signame Geodatabase Table 
#   Analysis is merely expanded to cover all training fields, not merely those misclassified (as the signature could still be contributing to the cumulative effect of mislabeling fields

training_label_sig_table_name = region_and_time_caps + '_training_label_signames_' + iteration_number
training_label_sig_table = os.path.join(gdb_path, training_label_sig_table_name)

TabulateArea(in_zone_data = training_label_shapefile, zone_field = 'crop', in_class_data = classified_tiff, class_field = 'Value', out_table = training_label_sig_table)

#----------------------------------------------------------------------------------------------

# 17. Generate Bad Signame Excel Workbook, a pivoted and manipulated version of Bad Signame Geodatabase Table

# Create numpy array from Training Label Signame Geodatabase Table
numpy_array_training_label_sig = arcpy.da.TableToNumPyArray(in_table = training_label_sig_table, field_names = '*')

# Create pandas data frame from numpy array
pandas_training_label_sig = pandas.DataFrame(data = numpy_array_training_label_sig)

# Replace column names from Bad Signame Geodatabase Table (derived from Reclassified Tiff attribute table, Value) to match corresponding Signame

# Create dictionary from Classified Tiff where key: value is Value: Class_Name
sig_dictionary = {}

# Populate empty dictionary key: value pairs using attribute table fields: Value and Class_Name from Classified Tiff

with arcpy.da.UpdateCursor(classified_tiff, ['Value', 'Class_Name']) as eCursor: 
    for row in eCursor:
        sig_dictionary[row[0]] = row[1]

# Add string Value_ to each key in dictionary so that dictionary reads old: new column names for data frame
sig_dictionary = {f'VALUE_{k}': v for k, v in sig_dictionary.items()}

for k, v in sig_dictionary.items():
    print(k,v)

# Replace column headers in data frame with Signame by using dictionary
pandas_training_label_sig.rename(columns = sig_dictionary, inplace = True) 

# Manipulate and pivot pandas data frame 

# Delete OBJECTID column
pandas_training_label_sig.drop('OBJECTID', axis = 1, inplace = True)

# Rename spurious column name CROP to true name Signame
pandas_training_label_sig = pandas_training_label_sig.rename({'CROP':'Signame'}, axis = 1)

# Reset index (row names) to values from CROP column
pandas_training_label_sig.set_index('Signame', inplace = True)

# Transpose data frame
pandas_training_label_sig = pandas_training_label_sig.transpose()

# Convert column names to string values
pandas_training_label_sig.columns = pandas_training_label_sig.columns.astype(str)
 
# Create list of column names
column_headers = list(pandas_training_label_sig.columns.values)

# Add summary column (i.e. total sum per row)
pandas_training_label_sig['Total_Area'] = pandas_training_label_sig[column_headers].sum(axis = 1) 

# Add Crop_label column, with value extracted from Signame

crop_label_list = pandas_training_label_sig.index.str.split('-').str[2].tolist() 
pandas_training_label_sig['Crop_label'] = pandas.to_numeric(crop_label_list)
pandas_training_label_sig['Crop_label'] = pandas_training_label_sig['Crop_label'].div(100)
pandas_training_label_sig['Crop_label'].fillna(-9999, inplace = True)
pandas_training_label_sig['Crop_label'] = pandas_training_label_sig['Crop_label'].astype(int) 
pandas_training_label_sig['Crop_label'] = pandas_training_label_sig['Crop_label'].astype(str) 

# Create two empty columns and set as numeric by filling with value of 0 instead of '' (which would create a series of string data type instead)
pandas_training_label_sig['Good'] = 0
pandas_training_label_sig['Bad'] = 0

# Create a new column that will be the normazized index (i.e. ranging from -1 to 1) of badness, where 1 is the worst (i.e. signature always classified incorrectly)
pandas_training_label_sig['Badness_Index'] = 0

# For each signature, aggregate area based on whether it was good (i.e. for the correct crop), or bad (i.e. classified area known to be another crop)

# Loop through each column, then loop through each row and if the Crop_label value for that row is the same as the column you're looping through, add that value to Good, otherwise add to Bad column
for h in column_headers:
    for i, row in pandas_training_label_sig.iterrows(): 
        if pandas_training_label_sig.at[i, 'Crop_label'] == h:
            pandas_training_label_sig.at[i, 'Good'] = pandas_training_label_sig.at[i, 'Good'] + pandas_training_label_sig.at[i, h]
        else:
            pandas_training_label_sig.at[i, 'Bad'] = pandas_training_label_sig.at[i, 'Bad'] + pandas_training_label_sig.at[i, h]

# Calculate badness index 
pandas_training_label_sig['Badness_Index'] = ((pandas_training_label_sig['Bad'] - pandas_training_label_sig['Good']) / (pandas_training_label_sig['Bad'] + pandas_training_label_sig['Good']))

# Create list of columns to total
columns_to_total = list(pandas_training_label_sig.columns.values)

columns_not_total = ['Crop_label', 'Badness_Index']
for i in columns_not_total:
    if i in columns_to_total:
        columns_to_total.remove(i)

# Add summary row (i.e. total sum per column) 
pandas_training_label_sig.loc['Total',:] = pandas_training_label_sig[columns_to_total].sum(axis = 0)

# Print text (which can be copied and pasted into ERDAS IMAGINE Signature Editor criteria search) of recommended signatures to delete 

badness_thresholds = [-0.5, -0.7, -1]

for b in badness_thresholds:
    list_signatures_remove = pandas_training_label_sig.index[pandas_training_label_sig['Badness_Index'] > b].tolist()
    list_signatures_remove = ['$"Signature Name" == "{}"'.format(i) for i in list_signatures_remove]
    string_signatures_remove = ' OR '.join(list_signatures_remove)
    arcpy.AddMessage('The following text can be copied into ERDAS IMAGINE Signature Editor Criteria box; it corresponds to Signames that had a Badness Index of ' + str(b) + ' greater than : ' + string_signatures_remove)

# 16.4 Create Excel file from pandas data frame 

bad_sig_excel_name = region_and_time_caps + '_Bad_Signame_Workbook.xlsx'
bad_sig_excel = os.path.join(docs_path, bad_sig_excel_name)
iteration_name = 'iteration_' + iteration_number

# Import function, load_workbook, from openpyxl which allows you to add a sheet to an already existing Excel file
from openpyxl import load_workbook 

# Define function returning boolean value of whether or not a file has a handle on it (Source: https://stackoverflow.com/questions/11114492/check-if-a-file-is-not-open-nor-being-used-by-another-process)
if os.path.isfile(bad_sig_excel): 

    def has_handle(file_path):
            for proc in psutil.process_iter():
                try:
                    for item in proc.open_files():
                        if file_path == item.path:
                            return True
                except Exception:
                    pass
        
            return False
                
# Test whether Bad Signame Excel Workbook has a handle on it (is open on the local computer) and pause the script to give the user a chance to close it so as to avoid PermissionError (below)
    my_test = has_handle(bad_sig_excel)
    if my_test == True:
        while my_test == True:
            arcpy.AddWarning('Please close ' + bad_sig_excel + ' in order to proceed')
            time.sleep(5)
            my_test = has_handle(bad_sig_excel)
        arcpy.AddMessage(bad_sig_excel + ' is now closed; continuing...')

# Once Bad Signame Excel Workbook has no handle on it, add a sheet for this iteration, or create the Excel file if one does not exist (in the case of first iteration)

    book = load_workbook(bad_sig_excel)
    writer = pandas.ExcelWriter(bad_sig_excel, engine ='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    pandas_training_label_sig.to_excel(writer, iteration_name)
    
    try:
        writer.save()
    except PermissionError:
        arcpy.AddError('Please close ' + bad_sig_excel + ' and re-run tool')
        sys.exit(0)       
    else:
        arcpy.AddMessage('Added new sheet in Bad Signatures Excel Workbook: ' + str(bad_sig_excel) + ' for iteration number: ' + str(iteration_number))

else:
    pandas_training_label_sig.to_excel(excel_writer = bad_sig_excel, sheet_name = iteration_name)
    arcpy.AddMessage('Generated Bad Signatures Excel Table: ' + str(bad_sig_excel))



